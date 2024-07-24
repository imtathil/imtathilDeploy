import pickle

#To solve GUI problems areeeeeeeeej 
import matplotlib
matplotlib.use('Agg')

#these libraries are used for the Model 
import warnings
from transformers import GPT2LMHeadModel, GPT2Tokenizer, TextDataset, DataCollatorForLanguageModeling, Trainer, TrainingArguments, AutoTokenizer
import torch
import re

#these libraries are used for the pdf here test test
import pandas as pd
from fpdf import FPDF
import arabic_reshaper
from bidi.algorithm import get_display
import subprocess


#Main Class (Can be called anywhere in the app)
class DataPreprocessor:
    def __init__(self, excel_file):
        self.excel_file = excel_file

    def preprocess(self):
        warnings.simplefilter(action='ignore', category=FutureWarning)
        pd.options.mode.chained_assignment = None 
        file = pd.ExcelFile(self.excel_file)
        sheets = [x for x in file.sheet_names if 'ECC' in x]

        def get_tables(sheet_name):
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            start_row = 5
            end_row = None
            dataframes = []

            for index, row in df.iterrows():
                if row.isnull().all():
                    if end_row is not None:
                        dataframes.append(df[start_row:end_row])
                    start_row = index + 2
                    end_row = None
                else:
                    end_row = index + 1

            if end_row is not None:
                dataframes.append(df[start_row:end_row])
            return dataframes

        dataframes = []
        for sheet in sheets:
            dataframes += get_tables(sheet)

        headers2 = ['رقم الضابط الأساسي',
                    'رقم الضابط الفرعي',
                    'تفاصيل الضابط',
                    'مستوى الالتزام الضابط الأساسي',
                    'مستوى الالتزام الضابط الفرعي',
                    'ملاحظات']

        new_dfs = []
        for df in dataframes[1:]:
            df.reset_index(drop=True, inplace=True)
            df = df.iloc[3:]
            df.drop(df.columns[4], axis=1, inplace=True)
            try:
                df.drop(3, inplace=True)
            except:
                continue
            df = df.iloc[:, :-2]
            df.columns = headers2
            new_dfs.append(df)

        finalDF = pd.concat(new_dfs, ignore_index=True)
        finalDF['إجراءات التصحيح المقترحة'] = None

        def rearrange_sentence(sentence):
            # Split the sentence into words
            words = sentence.split()

            # Generate the order dynamically (reverse order)
            order = list(range(len(words)-1, -1, -1))

            # Rearrange the words according to the specified order
            rearranged_sentence = " ".join(words[i] for i in order)

            return rearranged_sentence
        
        model_name = r"blog\static\blog\model\TrainedModel" #Traiend model path
        tokenizer = GPT2Tokenizer.from_pretrained(model_name)
        model = GPT2LMHeadModel.from_pretrained(model_name)
        tokenizer.pad_token_id = tokenizer.eos_token_id

        def generate_text(prompt="", max_length=200, num_return_sequences=1, no_repeat_ngram_size=2, beam_size=6):
            input_ids = tokenizer.encode(prompt, return_tensors="pt")
            attention_mask = torch.ne(input_ids, tokenizer.pad_token_id)

            output = model.generate(input_ids,
                            max_length=max_length,
                            num_return_sequences=num_return_sequences,
                            no_repeat_ngram_size=no_repeat_ngram_size,
                            attention_mask=attention_mask,
                            pad_token_id=model.config.eos_token_id,
                            do_sample=True,
                            early_stopping=True,
                            num_beams=beam_size,
                            temperature=0.1 )
            
            # Decode the generated text
            generated_texts = []
            
            for i in range(output.shape[0]):
                generated_text = tokenizer.decode(output[i], skip_special_tokens=True)

                #post-process
                generated_text = generated_text.strip()
                generated_text = re.sub(r'\s+', ' ', generated_text)
                generated_text = re.sub(r'[^\w\s]', '', generated_text)
                generated_texts.append(generated_text)

            return generated_texts

        for index, row in finalDF.iterrows():
                if (row['مستوى الالتزام الضابط الأساسي'] in ["غير مطبق  - Not Implemented", "مطبق جزئيًا  - Partially Implemented", "لاينطبق على الجهة  - Not Applicable"] or
                    row['مستوى الالتزام الضابط الفرعي'] in ["غير مطبق  - Not Implemented", "مطبق جزئيًا  - Partially Implemented", "لاينطبق على الجهة  - Not Applicable"]):
                    input_text = row['تفاصيل الضابط']
                    generated_text = generate_text(input_text)
                    generated_text_str = ' '.join(generated_text)
                    finalDF.at[index, 'إجراءات التصحيح المقترحة'] = rearrange_sentence(generated_text_str)
                    #print(generated_text_str)
                    print(rearrange_sentence(generated_text_str))
                else: finalDF.at[index, 'إجراءات التصحيح المقترحة'] = "لا يوجد مقترح"

        #The following part is to make the dataframe more represenable (by filling the NaN fields)
        finalDF['ملاحظات'] = finalDF['ملاحظات'].fillna('_')
        finalDF['رقم الضابط الأساسي'] = finalDF['رقم الضابط الأساسي'].fillna('_')

        #dropped this column cause there is no space in the pdf file created 
        finalDF = finalDF.drop(columns=['تفاصيل الضابط', 'ملاحظات'])


#To calculaet the percentage of compliance نسبة مستوى الالتزام الضوابط في آداة الهيئة
#total number of main controls 
        numofControls  = 0
        #num of controls that are FULLY implemented 
        implementedControls = 0
        #num of controls that are PARTIALLY implemented
        partiallyImplementedControls = 0 
        #num of controls that are NOT implemented
        notImplementedControls=0
        #num of controls thar are NOT APPLICAPLE 
        notApplicapleControls=0

        for index, row in finalDF.iterrows():
            if (row['مستوى الالتزام الضابط الأساسي'] in ["مطبق كليًا  - Implemented"]):
                numofControls  += 1
                implementedControls += 1
            elif (row['مستوى الالتزام الضابط الأساسي'] in ["مطبق جزئيًا  - Partially Implemented"]):
                numofControls  += 1
                partiallyImplementedControls += 1
            elif (row['مستوى الالتزام الضابط الأساسي'] in ["غير مطبق  - Not Implemented"]):
                numofControls  += 1
                notImplementedControls += 1
            elif (row['مستوى الالتزام الضابط الأساسي'] in ["لاينطبق على الجهة  - Not Applicable"]):
                numofControls  += 1 
                notApplicapleControls += 1
            else: 
                continue

        print("عدد الضوابط الأساسية:", numofControls)
        print("عدد الضوابط الأساسية المطبقة كلياً:", implementedControls)
        print("عدد الضوابط الأساسية المطبقة جزيئاَ:", partiallyImplementedControls)
        print("عدد الضوابط الأساسية الغير مطبقة:", notImplementedControls)
        print("عدد الضوابط الأساسية التي تنطبق لا على الجهة:", notApplicapleControls)


        #calculation of the TOTAL compliance percentage
        totalCompliancePercentage = ((numofControls - (partiallyImplementedControls+notImplementedControls+notApplicapleControls)) / numofControls ) *100
        rounded_totalCompliancePercentage = round(totalCompliancePercentage, 1)
        print( "نسبة التزام الضوابط الأساسية:", "%", rounded_totalCompliancePercentage)

        #calculation of the PARTIALLY implemented percentage
        totalPartiallyImplementedControls = ((numofControls - (implementedControls+notImplementedControls+notApplicapleControls)) / numofControls ) *100
        rounded_totalPartiallyImplementedControls = round(totalPartiallyImplementedControls, 1)
        print( "نسبة الضوابط الأساسية المطبقة جزئيًا: ", "%", rounded_totalPartiallyImplementedControls)

        #calculation of the Not implemented percentage
        totalNotImplementedControls = ((numofControls - (implementedControls+partiallyImplementedControls+notApplicapleControls)) / numofControls ) *100
        rounded_totalNotImplementedControls  = round(totalNotImplementedControls, 1)
        print( "نسبة الضوابط الأساسية الغير مطبقة: ", "%", rounded_totalNotImplementedControls)

        #calculation of the Not implemented percentage
        totalNotApplicapleControls = ((numofControls - (implementedControls+partiallyImplementedControls+notImplementedControls)) / numofControls ) *100
        rounded_totalNotApplicapleControls  = round(totalNotApplicapleControls, 1)
        print( "نسبة الضوابط الأساسية التي لا تنطبق على الجهة:", "%", rounded_totalNotApplicapleControls)



        
       ## Control's Table PDF Code, this is the output fter uploading the file and using generative AI! ## 

        from reportlab.pdfgen import canvas
        from reportlab.lib import utils
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase import pdfmetrics
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
        import openpyxl
        import arabic_reshaper
        from bidi.algorithm import get_display
        from PyPDF2 import PdfReader, PdfWriter
        import numpy as np  # For numerical operations
        from reportlab.lib.units import inch
        import subprocess
        from fpdf import FPDF
        from bidi.algorithm import get_display
        import subprocess
        
        ## Code To generate نسبة مستوى الالتزام الضوابط في أداة الهيئة as a pdf ##
        # Register the custom font file
        pdfmetrics.registerFont(TTFont('ArabicFont', r"blog\static\blog\Fonts\Font.ttf"))  # Replace 'path_to_your_arabic_font.ttf' with the actual file path

        # Create a PDF document
        doc = SimpleDocTemplate("compliance_Stats.pdf", pagesize=letter)

        elements = []

        # Reshape the title using Arabic reshaper
        title_text = "احصائيات عامة عن مدى التزام الجهة:"
        title_text_reshaped = arabic_reshaper.reshape(title_text)
        title_text_display = get_display(title_text_reshaped)

        # Add title as a paragraph
        title_style = ParagraphStyle(name='TitleStyle', fontName='ArabicFont', fontSize=19, alignment=1, spaceBefore=30, spaceAfter=40)
        title = Paragraph(title_text_display, title_style)
        elements.append(title)


        # Reshape Arabic strings
        N_Controls = get_display(arabic_reshaper.reshape("عدد الضوابط الأساسية:"))
        N_implemented_Controls = get_display(arabic_reshaper.reshape("عدد الضوابط الأساسية المطبقة كلياً:"))
        N_implemented_Controls_Percentage = get_display(arabic_reshaper.reshape("نسبة التزام الضوابط الأساسية المطبقة كلياً:"))
        N_partially_Implemented_Controls_Percentage = get_display(arabic_reshaper.reshape("نسبة الضوابط الأساسية المطبقة جزئيًا:"))
        N_partially_Implemented_Controls = get_display(arabic_reshaper.reshape("عدد الضوابط الأساسية المطبقة جزئياً:"))
        N_notImplemented_Controls_P = get_display(arabic_reshaper.reshape("نسبة الضوابط الأساسية الغير مطبقة:"))
        N_notImplemented_Controls = get_display(arabic_reshaper.reshape("عدد الضوابط الأساسية الغير مطبقة:"))
        N_notApplicable_Controls_P = get_display(arabic_reshaper.reshape("نسبة الضوابط الأساسية التي لا تنطبق على الجهة:"))
        N_notApplicable_Controls = get_display(arabic_reshaper.reshape("عدد الضوابط الأساسية التي لا تنطبق على الجهة:"))

        # Data for the table (excluding the first cell in the header)
        data = [
            [numofControls, N_Controls],
            [implementedControls, N_implemented_Controls],
            [f"%{rounded_totalCompliancePercentage}", N_implemented_Controls_Percentage],
            [f"%{rounded_totalPartiallyImplementedControls}", N_partially_Implemented_Controls_Percentage],
            [partiallyImplementedControls, N_partially_Implemented_Controls],
            [f"%{rounded_totalNotImplementedControls}", N_notImplemented_Controls_P],
            [notImplementedControls, N_notImplemented_Controls],
            [f"%{rounded_totalNotApplicapleControls}", N_notApplicable_Controls_P],
            [notApplicapleControls, N_notApplicable_Controls],
        ]

        # Create a table
        table = Table(data)

        # Add style to the table
        style = TableStyle([
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CBE54E')),  # Header background color
            ('FONTNAME', (0, 0), (-1, -1), 'ArabicFont'),  # Use the registered font name
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])

        table.setStyle(style)

        # Add table to the PDF document
        elements.append(table)

        # Build the PDF document
        doc.build(elements)


        ## pie and bar charts for the 'مستوى الالتزام الضابط الأساسي' column (numbers) ##

        # finalDF is the DataFrame and 'مستوى الالتزام الضابط الأساسي' is the column of interest

        # Get all possible states
        all_states = finalDF['مستوى الالتزام الضابط الأساسي'].unique()

        # occurrences count for 'مستوى الالتزام الضابط الأساسي' column 
        occurrance_counts = finalDF['مستوى الالتزام الضابط الأساسي'].value_counts()

        # Reindex with all possible states and fill missing values with zeros
        occurrance_counts = occurrance_counts.reindex(all_states, fill_value=0)

        # Create a dictionary to store reshaped and displayed versions of states
        states_display = {state: get_display(arabic_reshaper.reshape(state)) for state in occurrance_counts.index}



        ## Control's Table PDF Code ##


        def arabic_text(text):
            """Reshape and get the display order of the Arabic text."""
            if isinstance(text, float):  # Convert float to string
                text = str(text)
            reshaped_text = arabic_reshaper.reshape(text)
            bidi_text = get_display(reshaped_text)
            return bidi_text

        def generate_pdf_with_table(data):
            pdf = FPDF(orientation='L')  # Set orientation to landscape
            pdf.add_font("Arial", style="", fname=r"blog\static\blog\Fonts\Font.ttf", uni=True) #the path of the font used
            pdf.add_page()
            pdf.set_font("Arial", size=10)
            col_widths = [pdf.get_string_width(arabic_text(header)) + 6 for header in data.columns]

            # Calculate column widths dynamically based on content
            for index, row in data.iterrows():
                for i, col in enumerate(row):
                    text_width = pdf.get_string_width(arabic_text(str(col))) + 6
                    col_widths[i] = max(col_widths[i], text_width)

            # Increase the width of the last column to ensure it's wide enough
            last_col_index = len(data.columns) - 1
            col_widths[last_col_index] = 100  # Set a fixed width for the last column

            # Calculate total width of the table
            total_table_width = sum(col_widths)

            # Calculate starting X position to center the table
            pdf_width = pdf.w - 2 * pdf.l_margin
            start_x = (pdf_width - total_table_width) / 2 + pdf.l_margin

            # Set the fill color for the header row
            header_fill_color = (203, 229, 78)  # Set the fill color for the header to #CBE54E

            # Create table headers
            pdf.set_fill_color(*header_fill_color)
            pdf.set_xy(start_x, pdf.get_y())
            for header, width in zip(data.columns, col_widths):
                pdf.cell(width, 10, arabic_text(header), border=1,  align='C', fill=True)
            pdf.ln()

            # Set the position for table rows
            pdf.set_y(pdf.get_y())
            pdf.set_text_shaping(use_shaping_engine=True, direction="ltr")
            # Create table rows
            for _, row in data.iterrows():
                fill_color = (240, 230, 140) if all(['لا يوجد مقترح' not in str(value) for value in row]) else (255, 255, 255)

                pdf.set_fill_color(*fill_color)
                pdf.set_x(start_x)
                for i, (col, width) in enumerate(zip(row, col_widths)):
                    text = arabic_text(str(col)) if not pd.isnull(col) else ""  # Check for empty cell
                    if i == last_col_index:
                        # Print the content of the last column as a multi-line
                        pdf.multi_cell(width, 10, str(col) if 'لا يوجد مقترح' in str(col) else rearrange_sentence(str(col)) , border=1, align='C', fill=True)
                    else:
                        # Print the content of other columns
                        if i ==0:
                            pdf.cell(width, 10, text , border=1,  align='C', fill=True)

                        else:
                            pdf.cell(width, 10, arabic_text(text) , border=1,  align='C', fill=True)

            pdf.output('Recommendation_Table.pdf', 'F')

        generate_pdf_with_table(finalDF)




        ## Read The Figuers in  "ملخص نتائج تقييم والتزام الجهة" as a dataframe ##



        #import pandas as pd
        import openpyxl
        from matplotlib.backends.backend_pdf import PdfPages

        def extract_data_from_excel(file_path, sheets_columns_rows):
            # Load the Excel file
            wb = openpyxl.load_workbook(file_path, data_only=True)  # Ensure data_only=True to get calculated values
            extracted_data_dict = {}

            for sheet_name, column, row_start, row_end in sheets_columns_rows:
                # Get the specific sheet
                sheet = wb[sheet_name]

                # Extract data for current sheet
                extracted_data = []

                # Iterate over the specified rows and columns
                for row_num in range(row_start, row_end + 1):
                    # Get the cell values from column B and C
                    cell_b = sheet['B' + str(row_num)].value
                    cell_c = sheet[column + str(row_num)].value if sheet[column + str(row_num)].value is not None else 0
                    extracted_data.append([cell_b, cell_c])  # Append the cell B value and calculated value of cell C

                # Create a DataFrame from the extracted data
                df = pd.DataFrame(extracted_data, columns=['حالة الضابط', 'عدد الضوابط'])
                df.set_index('حالة الضابط', inplace=True)  # Set 'Row' column as index
                extracted_data_dict[( column, row_start, row_end)] = df

            return extracted_data_dict

        # Function to create pie chart for each DataFrame
        def create_pie_chart(dataframe):
            plt.figure(figsize=(8, 6))
            plt.pie(dataframe['عدد الضوابط'], labels=dataframe.index, autopct='%1.1f%%', startangle=140)
            plt.title("Pie Chart")
            plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
            plt.show()

        file_path = self.excel_file
        sheet_Name =  "ملخص نتائج تقييم والتزام الجهة" 
        sheets_columns_rows = [(sheet_Name, "C", 7, 10), (sheet_Name, "C", 20, 23), (sheet_Name, "C", 43, 46), (sheet_Name, "C", 66, 69), (sheet_Name, "C", 89, 92), (sheet_Name, "C", 112, 115)]  # Specify the sheets, columns, and rows
        extracted_data_dict = extract_data_from_excel(file_path, sheets_columns_rows)



        ## Charts' PDF Code ##





        def extract_data_from_excel(file_path, sheets_columns_rows):
            # Load the Excel file
            wb = openpyxl.load_workbook(file_path, data_only=True)  # Ensure data_only=True to get calculated values
            extracted_data_dict = {}

            for sheet_name, column, row_start, row_end in sheets_columns_rows:
                # Get the specific sheet
                sheet = wb[sheet_name]

                # Extract data for current sheet
                extracted_data = []

                # Iterate over the specified rows and columns
                for row_num in range(row_start, row_end + 1):
                    # Get the cell values from column B and C
                    cell_b = sheet['B' + str(row_num)].value
                    cell_c = sheet[column + str(row_num)].value if sheet[column + str(row_num)].value is not None else 0
                    extracted_data.append([cell_b, cell_c])  # Append the cell B value and calculated value of cell C

                # Create a DataFrame from the extracted data
                df = pd.DataFrame(extracted_data, columns=['حالة الضابط', 'عدد الضوابط'])
                df.set_index('حالة الضابط', inplace=True)  # Set 'Row' column as index
                extracted_data_dict[(sheet_name, column, row_start, row_end)] = df

            return extracted_data_dict

        # Function to create pie chart for each DataFrame
        def create_pie_chart(dataframe, title):
            # Reshape labels
            labels = [get_display(arabic_reshaper.reshape(label)) for label in dataframe.index]
            
            plt.pie(dataframe['عدد الضوابط'], labels=labels, autopct='%1.1f%%', startangle=140)
            plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.


        # Function to create table for each DataFrame
        def create_table(dataframe, table_size=(2.5, 3.5)):
            ax = plt.gca()
            ax.axis('off')
            
            # Reshape column labels
            col_labels = [get_display(arabic_reshaper.reshape(label)) for label in dataframe.columns]
            
            # Reshape row labels
            row_labels = [get_display(arabic_reshaper.reshape(label)) for label in dataframe.index]

            table = ax.table(cellText=dataframe.values, colLabels=col_labels, rowLabels=row_labels, loc='center', cellLoc='center')
            table.auto_set_font_size(False)
            table.set_fontsize(12)


        # Example usage
        ## **** change the path as it should be taken from the user in the website 
        file_path = self.excel_file
        sheet_Name =  "ملخص نتائج تقييم والتزام الجهة" 
        sheets_columns_rows = [(sheet_Name, "C", 7, 10), (sheet_Name, "C", 20, 23), (sheet_Name, "C", 43, 46), (sheet_Name, "C", 66, 69), (sheet_Name, "C", 89, 92), (sheet_Name, "C", 112, 115)]  # Specify the sheets, columns, and rows
        extracted_data_dict = extract_data_from_excel(file_path, sheets_columns_rows)

        # Specify titles manually
        titles = ["المستوى العام لتقييم الأمن السيبراني للجهة", "حوكمة الأمن السيبراني", "تعزيز الأمن السيبراني", "صمود الأمن السيبراني", "الأمن السيبراني المتعلق بالأطراف الخارجية والحوسبة السحابية", "الأمن السيبراني لأنظمة التحكم الصناعي"]

        # Create a PDF file for storing pie charts and tables
        pdf_file = "Summary_Charts.pdf"

        # Create PDF document
        with PdfPages(pdf_file) as pdf:
            # Visualize each DataFrame as a pie chart and table and save to PDF
            for (key, df), title in zip(extracted_data_dict.items(), titles):
                # Create a new figure for each page
                plt.figure(figsize=(12, 6))
                
                # Add centered title for the page
                plt.suptitle(get_display(arabic_reshaper.reshape(title)), fontsize=16)
                
                # Create pie chart subplot
                plt.subplot(1, 2, 1)
                create_pie_chart(df, title)
                
                # Create table subplot
                plt.subplot(1, 2, 2)
                create_table(df, title)
                
                # Adjust layout
                plt.tight_layout(rect=[0, 0.03, 1, 0.95])  # Add a small space at the top for the title
                
                # Save figure to PDF
                pdf.savefig()
                plt.close()



        ## Code for  'معلومات أساسية عن الجهة'  Page ##
        # Function to extract data from an Excel sheet
        def extract_data_from_excel(excel_file, sheet_name, skip_rows=0):
            try:
                # Read data from Excel file. 
                # - excel_file: the path to the Excel file.
                # - sheet_name: the name of the sheet to read.
                # - skip_rows: number of rows to skip before reading data (0-based).
                df = pd.read_excel(excel_file, sheet_name=sheet_name, skiprows=skip_rows + 1, header=None)
                # Drop rows where all elements are NaN
                df = df.dropna(how='all')
                return df
            except Exception as e:
                # Print error message if reading fails
                print("Error occurred while reading data:", e)
                return None

        # Function to create a PDF containing tables from the extracted data
        def create_pdf_table(data, pdf_file):
                # Fill NaN values with empty strings for better table formatting
                data = data.fillna('')


                # Use PdfPages to create a PDF file
                with PdfPages(pdf_file) as pdf:
                    # Group the data by the first column
                    for idx, group in data.groupby(data.columns[0]):
                        # If the group is not empty after dropping the first column
                        if not group.drop(columns=[data.columns[0]]).empty:
                            # Convert the group data to a list of lists for the table
                            table_data = group.drop(columns=[data.columns[0]]).applymap(str).values.tolist()

                            # Create a figure and axis for the table
                            fig, ax = plt.subplots(figsize=(10, 5))
                            # Hide the axis
                            ax.axis('tight')
                            ax.axis('off')
                            # Create the table, flipping the data horizontally
                            table = ax.table(cellText=np.flip(table_data, axis=1), loc='center', cellLoc='center')

                            # Set font size manually
                            table.auto_set_font_size(False)
                            table.set_fontsize(10)

                            # Process each cell in the table
                            for cell in table.get_celld().values():
                                cell_text = cell.get_text()
                                # Reshape the Arabic text for correct display
                                reshaped_text = arabic_reshaper.reshape(cell_text.get_text())
                                # Apply bidirectional algorithm to the reshaped text
                                bidi_text = get_display(reshaped_text)
                                cell_text.set_text(bidi_text)
                                cell_text.set_fontsize(10)
                                cell_text.set_fontfamily('Arial')

                            # Adjust column widths
                            for col in range(len(group.drop(columns=[data.columns[0]]).columns)):
                                col_width = max([len(str(group.drop(columns=[data.columns[0]]).iloc[row, col])) for row in range(len(group.drop(columns=[data.columns[0]])))])
                                table.auto_set_column_width(col)
                                
                            # Adjusting cells for specified rows
                            for (i, j), cell in table.get_celld().items():
                                text = cell.get_text().get_text()
                                num_lines = len(text.split('\n'))
                                cell.set_width(col_width * 0.1)
                                if i in [1, 5, 9]:  # Example rows to highlight
                                    cell.set_facecolor("#CBE54E")
                                    if j == 3:
                                        # Merge cells in the specified row and column
                                        merged_text = "".join([table_data[i][k] for k in range(len(table_data[i]))])
                                        reshaped_text = arabic_reshaper.reshape(merged_text)
                                        bidi_text = get_display(reshaped_text)
                                        cell.get_text().set_text(bidi_text)
                                        cell.set_text_props(ha='right')
                                        cell.set_width(1)
                                    else:
                                        cell.set_visible(False)
                                if i in [10, 11, 12, 13]:  # Example rows for specific height adjustment
                                    cell.set_height(0.12)
                                elif i == 14:
                                    cell.set_height(0.2)
                                else:
                                    cell.set_height(num_lines * 0.1)

                            # Save the current figure to the PDF
                            pdf.savefig(fig, bbox_inches='tight')
                            plt.close()

        # Define the Excel file, sheet name, and output PDF file
        ## **** change the path as it should be taken from the user in the website **** ##
        excel_file = self.excel_file # Path to the Excel file
        sheet_name = 'معلومات أساسية عن الجهة'  # Name of the sheet to read
        pdf_file = "Organizations_Data.pdf" # Output PDF file name

        # Extract data from the specified Excel sheet
        data = extract_data_from_excel(excel_file, sheet_name, skip_rows=3)
        create_pdf_table(data,pdf_file)





### Cover Page Code ###






        ### Code to Generate the Total PDF File With All the Content ###
        def create_pdf(title, text, image_path1, image_path2, output_file, font_path):

            # Register custom font
            pdfmetrics.registerFont(TTFont('CustomFont', font_path))

            # Create a canvas
            c = canvas.Canvas(output_file, pagesize=letter)

            # Get page width and height
            page_width, page_height = letter


            #inch = 1
            # Set text on the first page
            c.setFont("CustomFont", 12)
            text_lines = text.split('\n')
            y_text = page_height - 1.5 * inch
            for line in text_lines:
                text_width = c.stringWidth(line, "CustomFont", 12)
                text_x = (page_width - text_width) / 2
                c.drawString(text_x, y_text, line)
                y_text -= 0.3 * inch

            # Add image2 on the first page at the end
            img2 = utils.ImageReader(image_path2)
            img2_width, img2_height = img2.getSize()
            aspect_ratio2 = img2_height / float(img2_width)
            img2_x = 10  # Align with the left side of the page
            img2_y = inch  # Set y position to inch from the bottom
            c.drawImage(image_path2, img2_x, img2_y, width=page_width, height=page_height)

            # Add image1 on the first page, on top of image2
            img = utils.ImageReader(image_path1)
            img_width, img_height = img.getSize()
            aspect_ratio = img_height / float(img_width)
            img_x = (page_width - 5*inch) / 2
            img_y = y_text - (5 * inch * aspect_ratio)
            c.drawImage(image_path1, img_x, img_y, width=5*inch, height=5*inch * aspect_ratio)

            # Set title on the first page
            title_width = c.stringWidth(title, "CustomFont", 16)
            title_x = (page_width - title_width) / 2
            c.setFont("CustomFont", 16)
            c.drawString(title_x, page_height - 10 * inch, title)

            # Save the PDF
            c.save()


        # Example usage
        title = get_display(arabic_reshaper.reshape("هذا التقرير تم إنشاؤه باستخدام أداة إمتثل"))
        text = get_display(arabic_reshaper.reshape("هذا تقرير تم إنشاؤه لأداة إمتثل"))
        image_path1 = r"blog\static\blog\images\ImtathilPDF.jpg"
        image_path2 = r"blog\static\blog\images\back.jpg"
        font_path = r"blog\static\blog\Fonts\Font.ttf"
        output_file = "Cover_Page.pdf"

        create_pdf(title, text, image_path1, image_path2, output_file, font_path)

        input_paths = ['Cover_Page.pdf','Organizations_Data.pdf', "compliance_Stats.pdf",'Recommendation_Table.pdf', 'Summary_Charts.pdf']

        return input_paths
    

    def mergepdf(self):
        import io
        from PyPDF2 import PdfWriter, PdfReader
        def merge_pdfs(input_paths):
            writer = PdfWriter()
            for path in input_paths:
                reader = PdfReader(path)
                for page_num in range(len(reader.pages)):
                    page = reader.pages[page_num]
                    writer.add_page(page)
            
            pdf_bytes = io.BytesIO()
            writer.write(pdf_bytes)
            pdf_bytes.seek(0)
            return pdf_bytes
        

        input_paths = ['Cover_Page.pdf','Organizations_Data.pdf', "compliance_Stats.pdf",'Recommendation_Table.pdf', 'Summary_Charts.pdf']

        pdf=merge_pdfs(input_paths)
        return pdf

        